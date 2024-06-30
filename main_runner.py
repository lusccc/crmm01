import json
import os
import random
from datetime import datetime

import numpy as np
import openpyxl
import pandas as pd
import torch
import torchinfo
from torch.utils.data import ConcatDataset
from transformers import AutoTokenizer, Trainer, EarlyStoppingCallback, HfArgumentParser, BertTokenizer
from transformers.utils import logging

from crmm import runner_setup
from crmm.arguments import BertModelArguments, MultimodalDataArguments, CrmmTrainingArguments
from crmm.dataset.multimodal_data import MultimodalData
from crmm.dataset.multimodal_dataset import MultimodalDatasetCollator
from crmm.metrics import calc_classification_metrics
from crmm.models.multi_modal_dbn import MultiModalModelConfig, MultiModalDBN, MultiModalForClassification, ModelForExplain
from explain_visual import lime_visual

logger = logging.get_logger('transformers')

os.environ['COMET_MODE'] = 'DISABLED'
os.environ["TOKENIZERS_PARALLELISM"] = "true"
os.environ["TRANSFORMERS_OFFLINE"] = "1"
# os.environ["WANDB_DISABLED"] = "true"

# 设置随机种子
seed = 3407
torch.manual_seed(seed)
np.random.seed(seed)
random.seed(seed)


def main(bert_model_args: BertModelArguments,
         data_args: MultimodalDataArguments,
         training_args: CrmmTrainingArguments):
    training_args.seed = seed
    task = training_args.task
    # @@@@ 1. TABULAR COLUMNS
    # note: the num and cat cols will be automatically inferred in `data.crmm_data.MultimodalData`
    label_col = data_args.label_col  # 'Rating'
    text_cols = data_args.text_cols.split(',')  # ['secText', 'secKeywords']
    label_list = np.load(os.path.join(data_args.data_path, 'label_list.npy'), allow_pickle=True)

    # @@@@ 2. DATASET
    mm_data = MultimodalData(data_args,
                             label_col=label_col,
                             label_list=label_list,
                             text_cols=text_cols,
                             num_transform_method=data_args.numerical_transformer_method)
    train_dataset, test_dataset, val_dataset = mm_data.get_datasets()
    n_labels = len(label_list)
    if task == 'pretrain':
        # if pretrain task, concat train and val to unsupervised train!
        train_dataset = ConcatDataset([train_dataset, val_dataset])
        val_dataset = None
    nunique_cat_nums, cat_emb_dims = mm_data.get_nunique_cat_nums_and_emb_dim(equal_dim=None)

    # @@@@ 3. MODEL
    bert_args = {'pretrained_model_name_or_path': bert_model_args.bert_model_name,
                 'cache_dir': bert_model_args.cache_dir,
                 'local_files_only': bert_model_args.load_hf_model_from_cache}
    tokenizer = AutoTokenizer.from_pretrained(**bert_args)
    default_model_config_params = dict(n_labels=n_labels,
                                       num_feat_dim=test_dataset.numerical_feats.shape[1],
                                       nunique_cat_nums=nunique_cat_nums,
                                       cat_emb_dims=cat_emb_dims,
                                       use_modality=training_args.use_modality,
                                       modality_fusion_method=training_args.modality_fusion_method,
                                       bert_args=bert_args,
                                       bert_model_name=bert_model_args.bert_model_name)
    if task == 'pretrain':
        model_config = MultiModalModelConfig(**default_model_config_params,
                                             use_hf_pretrained_bert=bert_model_args.use_hf_pretrained_bert_in_pretrain,
                                             freeze_bert_params=bert_model_args.freeze_bert_params,
                                             pretrained=False)
        model = MultiModalDBN(model_config)
    elif task == 'fine_tune':
        # finetune, model load from saved dir
        model_config = MultiModalModelConfig.from_pretrained(training_args.pretrained_model_dir)
        model_config.use_hf_pretrained_bert = False
        model = MultiModalForClassification.from_pretrained(training_args.pretrained_model_dir, config=model_config)
    elif task == 'fine_tune_from_scratch':
        model_config = MultiModalModelConfig(**default_model_config_params,
                                             use_hf_pretrained_bert=False,
                                             freeze_bert_params=bert_model_args.freeze_bert_params,
                                             pretrained=True)  # manually set pretrained to True!
        # create model, where the bert model is loaded from hf model in models.rbm_factory.RBMFactory._get_bert
        model = MultiModalForClassification(model_config)
    elif task == 'explain_visual':
        model_config = MultiModalModelConfig.from_pretrained(training_args.pretrained_model_dir)
        model_config.use_hf_pretrained_bert = False
        model = ModelForExplain.from_pretrained(training_args.pretrained_model_dir, config=model_config)
    else:
        raise ValueError(f'Unknown task: {task}')
    logger.info(f'\n{model}')
    logger.info(f'\n{torchinfo.summary(model, verbose=0)}')

    ####### SHAP VISUAL #######
    """
    python main_runner.py --task shap_visual --bert_model_name prajjwal1/bert-tiny --use_hf_pretrained_bert_in_pretrain true --freeze_bert_params false --use_modality text --modality_fusion_method conv --text_cols secKeywords --per_device_train_batch_size 300 --num_train_epochs 200 --patience 1000 --dataset_name cr --dataset_info  --data_path ./data/cr_cls2_mixed_st10_kw20 --output_dir ./exps/fine_tune_2024-06-13_19-37-43_UMY/output --logging_dir ./exps/fine_tune_2024-06-13_19-37-43_UMY/logging --pretrained_model_dir ./exps/pretrain_2024-06-13_19-37-43_Mr3/output --save_excel_path ./excel/cr_cls2_modality_txtcol(secKeywords)_(0621_aft_pre10).xlsx
    """
    if task == 'explain_visual':
        model.to(training_args.device)
        model.eval()
        # shap_visual(model, tokenizer, test_dataset)
        lime_visual(model, tokenizer, test_dataset)
        return

    # @@@@ 4. TRAINING
    get_trainer = lambda model: Trainer(
        model=model,
        args=training_args,
        train_dataset=train_dataset,
        eval_dataset=val_dataset,  # can be None in pretrain
        compute_metrics=calc_classification_metrics if 'fine_tune' in task else None,
        callbacks=[EarlyStoppingCallback(training_args.patience)] if 'fine_tune' in task else None,
        data_collator=MultimodalDatasetCollator(tokenizer, bert_model_args.max_seq_length)
    )
    if task == 'pretrain':
        # 1. pretrain each rbm of each modality
        model.set_pretrain_target('modality_rbm')
        trainer = get_trainer(model)
        trainer.train()
        # 2. if more than 1 modality, continue to pretrain joint rbm
        if len(model_config.use_modality) > 1:
            model.set_pretrain_target('joint_rbm')
            trainer.train()
        # should set pretrained to True. Hence, we can later continue to finetune the model
        model.config.pretrained = True
        trainer.save_model()
    else:
        trainer = get_trainer(model)
        trainer.train()
        trainer.save_model()

    best_model_checkpoint = trainer.state.best_model_checkpoint
    best_step = int(best_model_checkpoint.split("-")[-1]) if best_model_checkpoint else None
    logger.info(f"Best model path: {best_model_checkpoint}")
    logger.info(f"Best metric value: {trainer.state.best_metric}")

    # @@@@ 5. EVALUATION
    if task != 'pretrain':
        val_best_results = trainer.evaluate(eval_dataset=val_dataset)
        test_results = trainer.evaluate(eval_dataset=test_dataset)
        if task == 'fine_tune':
            with open(os.path.join(training_args.pretrained_model_dir, 'training_arguments.json')) as file:
                training_arguments = json.load(file)
            pretrain_batch_size = training_arguments['per_device_train_batch_size']
            pretrain_epoch = training_arguments['num_train_epochs']
        else:
            pretrain_step = None
            pretrain_batch_size = None
            pretrain_epoch = None

        basic_info = {
            'dataset': data_args.dataset_name,
            'dataset_info': data_args.dataset_info,
            'data_path': data_args.data_path,
            'bert_model': bert_model_args.bert_model_name,
            'numerical': 'num' in training_args.use_modality,
            'category': 'cat' in training_args.use_modality,
            'text': 'text' in training_args.use_modality,
            'pretrain_batch_size': pretrain_batch_size,
            'pretrain_epoch': pretrain_epoch,
            'fine_tune_best_step': best_step,
            'fine_tune_batch_size': training_args.per_device_train_batch_size,
            'fine_tune_epoch': training_args.num_train_epochs,
        }

        save_excel(val_best_results, test_results, basic_info, training_args.save_excel_path)


def save_excel(val_best_results, test_results, basic_info, excel_path):
    logger.info(f'** save results to {excel_path}')
    val_data = {f'val_{k}': v for k, v in val_best_results.items()}
    test_data = {f'test_{k}': v for k, v in test_results.items()}

    if os.path.exists(excel_path):
        book = openpyxl.load_workbook(excel_path)
        writer = pd.ExcelWriter(excel_path, engine='openpyxl')
        writer.book = book

        if 'Sheet1' in book.sheetnames:
            startrow = writer.sheets['Sheet1'].max_row
        else:
            startrow = 0

    else:
        writer = pd.ExcelWriter(excel_path, engine='openpyxl')
        startrow = 0

    data = {**{'Timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')},
            **basic_info, **val_data, **test_data, }
    df = pd.DataFrame([data])

    df.to_excel(writer, index=False, header=(startrow == 0), startrow=startrow)

    writer.save()
    writer.close()


if __name__ == '__main__':
    parser = HfArgumentParser([BertModelArguments, MultimodalDataArguments, CrmmTrainingArguments])
    _bert_model_args, _data_args, _training_args = parser.parse_args_into_dataclasses()
    _training_args = runner_setup.setup(_training_args, _data_args, _bert_model_args)
    logger.info(f'training_args: {_training_args}')

    main(_bert_model_args, _data_args, _training_args)
